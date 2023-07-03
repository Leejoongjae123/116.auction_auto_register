import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication, QTreeView, QFileSystemModel, QVBoxLayout, QPushButton, QInputDialog, \
    QLineEdit, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime, date, timedelta
import numpy
import datetime
from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from selenium.webdriver.common.alert import Alert


def chrome_browser(url):
    chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]  # 크롬 버전을 확인한다.
    driver_path = f'./{chrome_ver}/chromedriver.exe'
    if os.path.exists(driver_path):
        print(f"chromedriver is installed: {driver_path}")  # 있는 버전을 쓴다.
    else:
        print(f"install the chrome driver(ver: {chrome_ver})")  # 크롬을 최신 버전으로 설치한다.
        chromedriver_autoinstaller.install(True)

    options = webdriver.ChromeOptions()  # 크롬 옵션을 추가한다.
    # options.add_argument('headless')
    options.add_experimental_option("detach", True)  # 크롬 안 꺼지는 옵션 추가
    options.add_experimental_option("excludeSwitches", ["enable-logging"])  # 크롬 안 꺼지는 옵션 추가

    browser = webdriver.Chrome(driver_path, options=options)  # 크롬 드라이버를 할당
    browser.get(url)
    browser.maximize_window()
    browser.implicitly_wait(3)
    return browser
def load_excel(fname):
    wb = openpyxl.load_workbook(fname,data_only=True)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    data_list = []
    for i in range(2, no_row + 1):
        name = ws.cell(row=i, column=6).value
        if name == "" or name == None:
            print('데이타 더 이상 없음')
            break
        data_list.append(name)
        if len(data_list)>500:
            break
    print('urlList:',data_list)
    return data_list


class Thread(QThread):
    cnt = 0
    user_signal = pyqtSignal(str)  # 사용자 정의 시그널 2 생성

    def __init__(self, parent):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.

    def run(self):

        ProductNo = '5392069018'
        fname = 'list.xlsx'
        nameList = load_excel(fname)

        urlLogin = 'https://login.11st.co.kr/auth/front/selleroffice/login.tmall?returnURL=https%3A%2F%2Fsoffice.11st.co.kr%2Fview%2Fmain'
        browser = chrome_browser(urlLogin)

        id = 'moonstylecar'
        pw = '8823sun##'

        inputId = browser.find_element(By.CSS_SELECTOR, '#loginName')
        inputId.send_keys(id)
        time.sleep(0.5)

        inputPw = browser.find_element(By.CSS_SELECTOR, '#passWord')
        inputPw.send_keys(pw)
        time.sleep(0.5)

        buttonLogin = browser.find_element(By.CSS_SELECTOR, '#loginbutton')
        browser.execute_script("arguments[0].click();", buttonLogin)  #
        browser.implicitly_wait(3)

        urlProuctList = 'https://soffice.11st.co.kr/view/8006'
        browser.get(urlProuctList)
        browser.implicitly_wait(3)
        firstFlag = True
        for nameElem in nameList:
            print("상품({}) 등록 중".format(nameElem))
            element = browser.find_element(By.CLASS_NAME, "content_iframe")  # 태그찾기
            browser.switch_to.frame(element)  # 프레임 이동

            if firstFlag == True:
                ActionChains(browser).send_keys(Keys.PAGE_UP).perform()
                inputProuctNo = browser.find_element(By.CSS_SELECTOR, '#prdNo')
                inputProuctNo.send_keys(ProductNo)
                time.sleep(0.5)
                buttonSearch = browser.find_element(By.CSS_SELECTOR, '#btnSearch')
                buttonSearch.click()
                browser.implicitly_wait(3)
                time.sleep(0.5)
                ActionChains(browser).send_keys(Keys.PAGE_DOWN).perform()
                browser.implicitly_wait(3)
                time.sleep(0.5)
                while True:
                    try:
                        selectBox = browser.find_element(By.CSS_SELECTOR, '#row0dvdataGrid > div:nth-child(1)')
                        selectBox.click()
                        print("체크완료")
                        browser.implicitly_wait(3)
                        time.sleep(0.5)
                        btnCopy = browser.find_element(By.CSS_SELECTOR,
                                                       '#ext-gen1019 > div.soWrap > div.sobunchW > div.grid_top_button > div > a:nth-child(21)')
                        btnCopy.click()
                        browser.implicitly_wait(3)
                        break
                    except:
                        print("에러발생")
                time.sleep(0.5)
            if firstFlag == False:
                browser.implicitly_wait(3)
                time.sleep(0.5)
                btnCopy = browser.find_element(By.CSS_SELECTOR,
                                               '#ext-gen1019 > div.soWrap > div.sobunchW > div.grid_top_button > div > a:nth-child(21)')
                btnCopy.click()
                browser.implicitly_wait(3)

            firstFlag = False
            browser.switch_to.window(browser.window_handles[-1])
            browser.implicitly_wait(3)
            time.sleep(0.5)

            browser.find_element(By.CSS_SELECTOR, '#prdNm').click()
            ActionChains(browser).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            time.sleep(0.5)
            ActionChains(browser).send_keys(Keys.DELETE).perform()
            time.sleep(0.5)
            browser.find_element(By.CSS_SELECTOR, '#prdNm').send_keys(nameElem)
            time.sleep(0.5)

            cleanCheck = browser.find_element(By.CSS_SELECTOR,
                                              '#frmMain > div:nth-child(133) > table > tbody > tr:nth-child(2) > td > div > button')
            ActionChains(browser).move_to_element(cleanCheck).perform()
            cleanCheck.click()
            time.sleep(0.5)

            uploadMovie = browser.find_element(By.CSS_SELECTOR, '#movie_a')
            ActionChains(browser).move_to_element(uploadMovie).perform()
            uploadMovie.click()
            browser.implicitly_wait(3)
            time.sleep(0.5)

            browser.switch_to.window(browser.window_handles[-1])
            browser.implicitly_wait(3)
            time.sleep(0.5)

            btnSearch = browser.find_element(By.CSS_SELECTOR,
                                             '#video > div > table > tbody > tr:nth-child(5) > td > div')
            btnSearch.click()
            browser.implicitly_wait(3)
            time.sleep(1)

            movieDir = r'D:\PythonProjects\116.auction_auto_register\movie.mp4'
            pyperclip.copy(movieDir)
            time.sleep(0.5)
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.5)
            pyautogui.hotkey('enter')
            time.sleep(0.5)
            btnConfirm = browser.find_element(By.CSS_SELECTOR, '#movieConfirm > span')
            btnConfirm.click()

            while True:
                try:
                    alert = Alert(browser)
                    print(alert.text)
                    if alert.text.find("성공하였습니다") >= 0:
                        alert.accept()
                        break
                except:
                    print("확인중...")
                time.sleep(0.5)

            browser.implicitly_wait(3)
            time.sleep(0.5)
            print("창갯수:", len(browser.window_handles))
            browser.switch_to.window(browser.window_handles[-1])

            buttonRegister = browser.find_element(By.CSS_SELECTOR, '#dvPrdRegUpdBtn > a')
            ActionChains(browser).move_to_element(buttonRegister).perform()
            buttonRegister.click()
            time.sleep(0.5)


            while True:
                element = browser.find_element(By.ID, "_ifrmRegProcess")  # 태그찾기
                browser.switch_to.frame(element)  # 프레임 이동
                status = browser.find_element(By.CLASS_NAME, 'bottom_btnW').text
                print('status:', status)
                if status.find("등록되었습니다.") >= 0:
                    print("등록완료")
                    browser.close()
                    browser.switch_to.window(browser.window_handles[0])
                    break

        text="작업완료"
        self.user_signal.emit(text)

    def stop(self):
        pass


class Example(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path = "C:"
        self.index = None
        self.setupUi(self)
        self.setSlot()
        self.show()
        QApplication.processEvents()
        self.

    def start(self):
        print('11')
        self.x = Thread(self)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()

    def slot1(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit.append(str(data1))

    def setSlot(self):
        pass

    def setIndex(self, index):
        pass

    def quit(self):
        QCoreApplication.instance().quit()


app = QApplication([])
ex = Example()
sys.exit(app.exec_())







